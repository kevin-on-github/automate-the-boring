import openpyxl
import os

workbook = openpyxl.load_workbook('myworkbook.xlsx')
# print(type(workbook))

sheet = workbook['Sheet1']

# print(sheet['A1'].value)
print(sheet.cell(row=1, column=1).value)

for i in range(1, 6):
    print(sheet.cell(row=i, column=1).value)